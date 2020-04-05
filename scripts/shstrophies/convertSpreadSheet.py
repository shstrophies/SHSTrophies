# pip install pandas
# pip install jupyter
# pip install openpyxl
# pip install xlrd
import glob
from io import BytesIO

import numpy
import requests
import pandas as pd
import re
import os


def append_df_to_excel(df, filename, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    print('append_df_to_excel: filename=' + str(filename) + ', sheet_name=' + str(sheet_name) + ', startrow=' + str(
        startrow) + ', truncate_sheet=' + str(truncate_sheet) + ', to_excel_kwargs=' + str(to_excel_kwargs))

    """
    Append a dataframe [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain dataframe.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing dataframe to Excel file
      to_excel_kwargs : arguments which will be passed to `dataframe.to_excel()`
                        [can be dictionary]

    Returns: None
    """
    from openpyxl import load_workbook

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()


def convertDataframe(basefilename, sheetname, dataframe, title, url, lastdata=None, exportfile='export.xlsx', **data):
    print(
        "convertDataframe " + " basefilename=" + basefilename + ", sheet=" + sheetname + ", title=" + title + ", url=" + url + ", lastdata=" + str(
            lastdata))
    for key, value in data.items(): print("{}={}".format(key, value))
    # print("frame=" + frame.to_string())
    if (not title):
        print("title is empty, skip")
        return

    if (title == 'Year'):
        return
    dfa = dataframe[['Year', title]].assign(title=title).replace(0, numpy.nan).replace(0.0, numpy.nan).copy()
    # print("dfa=" + dfa.to_string())
    # fill any Nan Year with the previous row's non-Nan Year value
    dfb = dfa['Year'].fillna(method='ffill')
    # print("dfb=" + dfb.to_string())
    dfa.loc[dfb.index, 'Year'] = dfb
    # print("dfa=" + dfa.to_string())
    dfc = dfa[['Year', title]].assign(title=title).dropna().copy()
    # print("dfc=" + dfc.to_string())
    dfc.rename(columns={title: 'Player_Name'}, inplace=True)
    # print("dfc=" + dfc.to_string())
    dfd = dfc[['Year', 'Player_Name']].assign(Trophy_Title=title).copy()
    dfe = dfd.replace({'Year': r'[/-].*'}, {'Year': ''}, regex=True)[['Year', 'Trophy_Title', 'Player_Name']].copy()

    dff = dfe[['Year', 'Trophy_Title', 'Player_Name']].assign(Trophy_Image_URI=url).copy()
    dfg = dff[['Year', 'Trophy_Title', 'Trophy_Image_URI', 'Player_Name']].copy()
    # print("dfg=" + "\n" + dfg.to_string())

    # remove row that has image URL
    searchfor = ['image URL', 'image', 'URL']
    dfg = dfg[~dfg.Year.str.contains('|'.join(searchfor))]

    # append sheetname to trophy title
    dfg['Trophy_Title'] = dfg['Trophy_Title'].astype(str) + ' ' + '[' + sheetname + ']'
    print("output dataframe dfg=" + "\n" + dfg.to_string())

    outputSheet = getOutputSheet(basefilename)
    print('outputSheet=' + outputSheet)
    startRow = None
    header = True
    if (basefilename in lastdata):
        header = False

    append_df_to_excel(dfg, exportfile, sheet_name=outputSheet, startrow=startRow, header=header, index=False)

    theFileName = str(r'export_' + basefilename.replace(" ", "_") + '_' + title.replace(" ", "_") + '.xlsx')
    # dfg.to_excel(theFileName, sheet_name=outputSheet, index=False, header=True)
    return dfg


def getOutputSheet(name, **data):
    print("getOutputSheet " + " name=" + name)
    for key, value in data.items(): print("{}={}".format(key, value))
    outputSheet = str(name)
    dictOfSports = getDictOfSports()
    if(outputSheet in dictOfSports):
        return outputSheet
    else:
        for key in dictOfSports.keys():
            if(outputSheet.lower().find(key.lower()) >= 0):
                outputSheet = key
                return outputSheet
    return outputSheet

def getDictOfSports():
    sportsList = [
        'Volleyball',
        'Football',
        'Field Hockey',
        'Tennis',
        'Cross Country',
        'Water Polo',
        'Golf',
        'Dance',
        'Basketball',
        'Cheer',
        'Soccer',
        'Wrestling',
        'Baseball',
        'Swimming',
        'Swimming and Diving',
        'Track',
        'Volleyball',
        'Lacrosse',
        'Softball',
        'Badminton'
    ]
    dictOfSports = {i: len(sportsList) for i in sportsList}
    return dictOfSports


def getGender(**data):
    for key, value in data.items(): print("{}={}".format(key, value))
    if 'basefile' in data: basefile = data.get('basefile')
    if 'sheet' in data: sheet = data.get('sheet')

    gender = ""
    pattern = '^(G\.\s|Girl|Girls|GV|GJV)'
    if (re.match(pattern, basefile) | re.match(pattern, sheet)): gender = 'G'

    pattern = '^(B\.\s|Boy|Boys|BV|BFS)'
    if (re.match(pattern, basefile) | re.match(pattern, sheet)): gender = 'B'
    return gender


def getImageUrl(df, column, defaultUrl, **data):
    print("getImageUrl " + str(data))
    for key, value in data.items(): print("{}={}".format(key, value))
    imageUrl = defaultUrl
    firstRow = df[column].iloc[0]
    theType = type(firstRow)
    if (type(firstRow) == str):
        # print("first row=" + firstRow + " " + ",type=" + str(type(firstRow)))
        theFirstRow = str(firstRow)
        pattern = '^(http|https)://'
        if (re.match(pattern, theFirstRow)):
            imageUrl = firstRow
    return imageUrl


def convertExcelSheet(basefilename, sheetname, dataframe, exportfile, lastdata, **data):
    print("convertExcelSheet " + str(data))
    for key, value in data.items(): print("{}={}".format(key, value))
    dataframe.columns = dataframe.columns.to_series().apply(lambda x: x.strip())
    columns = dataframe.columns.values
    newCol = 'Year'
    print("renaming first column to " + newCol)
    dataframe.rename(columns={columns[0]: newCol}, inplace=True)
    dataframe.columns = dataframe.columns.to_series().apply(lambda x: x.strip())
    columns = dataframe.columns.values
    for column in columns:
        print("column=" + column)
        if ((column == 'Unnamed: 0')):
            newCol = 'Year'
            print('renaming column=' + column + " to " + newCol)
            dataframe.rename(columns={'Unnamed: 0': newCol}, inplace=True)
            column = 'Year'
        if (column != 'Year'):
            imageUrl = getImageUrl(dataframe, column, 'DEFAULT IMAGE', **data)
            print("imageUrl=" + imageUrl)
            convertDataframe(basefilename, sheetname, dataframe, column, imageUrl, lastdata, exportfile, **data)
            if (basefilename in lastdata):
                lastdata[basefilename].append(column)
            else:
                list = [column]
                lastdata[basefilename] = list
            print("last=" + str(lastdata))


def convertExcelFile(path, basefilename, exportpath, **data):
    print("convertExcelFile " + str(data))
    for key, value in data.items(): print("{}={}".format(key, value))
    lastdata = {}
    # df = pd.read_excel(excelFile)
    xls = pd.ExcelFile(path)
    for sheetName in xls.sheet_names:
        print("*******input sheet=" + sheetName)
        dataframe = xls.parse(sheetName)
        convertExcelSheet(basefilename, sheetName, dataframe, exportpath, lastdata, basefile=basefilename,
                          sheet=sheetName, **data)


def findFilesInFolder(path, pathList, extension, subFolders=True):
    """  Recursive function to find all files of an extension type in a folder (and optionally in all subfolders too)

    path:        Base directory to find files
    pathList:    A list that stores all paths
    extension:   File extension to find
    subFolders:  Bool.  If True, find files in all subfolders under path. If False, only searches files in the specified folder
    """

    try:  # Trapping a OSError:  File permissions problem I believe
        for entry in os.scandir(path):
            if entry.is_file() and entry.path.endswith(extension):
                pathList.append(entry.path)
            elif entry.is_dir() and subFolders:  # if its a directory, then repeat process as a nested function
                pathList = findFilesInFolder(entry.path, pathList, extension, subFolders)
    except OSError:
        print('Cannot access ' + path + '. Probably a permissions error')

    return pathList


extension = ".xlsx"
exportfile = 'export' + extension
os.remove(exportfile) if os.path.exists(exportfile) else None

# to test
# convertExcelFile('./data/Mr Torren_s initial data input/Fall Awards/G. Golf.xlsx', 'G. Golf', exportfile)
# convertExcelFile('./data/Mr Torren_s initial data input/Spring Awards/BVolleyball.xlsx', 'BVolleyball', exportfile)
# exit(0)


dir_name = r'./data'
pathList = []
pathList = findFilesInFolder(dir_name, pathList, extension, True)
print("all files=" + str(pathList))
for file in pathList:
    basename = os.path.basename(str(file))
    info = os.path.splitext(basename)
    baseFileName = info[0]
    print('************************************************************')
    print('************************************************************')
    print('************************************************************')
    print("file=" + str(file) + ", baseFileName=" + baseFileName)
    print('************************************************************')
    print('************************************************************')
    print('************************************************************')
    convertExcelFile(str(file), baseFileName, exportfile, file=str(file))
