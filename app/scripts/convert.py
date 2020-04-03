# pip install pandas
# pip install jupyter
# pip install openpyxl
# pip install xlrd
from io import BytesIO
import requests
import pandas as pd
import re

def append_df_to_excel(df, filename, sheet_name='Sheet1', startrow=None,
                           truncate_sheet=False,
                           **to_excel_kwargs):
        """
        Append a DataFrame [df] to existing Excel file [filename]
        into [sheet_name] Sheet.
        If [filename] doesn't exist, then this function will create it.

        Parameters:
          filename : File path or existing ExcelWriter
                     (Example: '/path/to/file.xlsx')
          df : dataframe to save to workbook
          sheet_name : Name of sheet which will contain DataFrame.
                       (default: 'Sheet1')
          startrow : upper left cell row to dump data frame.
                     Per default (startrow=None) calculate the last row
                     in the existing DF and write to the next row...
          truncate_sheet : truncate (remove and recreate) [sheet_name]
                           before writing DataFrame to Excel file
          to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                            [can be dictionary]

        Returns: None
        """
        from openpyxl import load_workbook

        # ignore [engine] parameter if it was passed
        if 'engine' in to_excel_kwargs:
            to_excel_kwargs.pop('engine')

        writer = pd.ExcelWriter(filename, engine='openpyxl')

        try:
            # try to open an existing workbook
            writer.book = load_workbook(filename)

            # get the last row in the existing Excel sheet
            # if it was not specified explicitly
            if startrow is None and sheet_name in writer.book.sheetnames:
                startrow = writer.book[sheet_name].max_row

            # truncate sheet
            if truncate_sheet and sheet_name in writer.book.sheetnames:
                # index of [sheet_name] sheet
                idx = writer.book.sheetnames.index(sheet_name)
                # remove [sheet_name]
                writer.book.remove(writer.book.worksheets[idx])
                # create an empty sheet [sheet_name] using old index
                writer.book.create_sheet(sheet_name, idx)

            # copy existing sheets
            writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
        except FileNotFoundError:
            # file does not exist yet, we will create it
            pass

        if startrow is None:
            startrow = 0

        # write out the new sheet
        df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

        # save the workbook
        writer.save()

def convertTable(sport, frame, theTitle, theUrl):
    print("convertTable sport=" + sport + ", title=" + theTitle + ", url=" + theUrl)
    if (theTitle == 'Year'):
        return
    dfc = frame[['Year', theTitle]].assign(title=theTitle).dropna().copy()
    dfc.rename(columns={theTitle: 'Player_Name'}, inplace=True)
    dfd = dfc[['Year', 'Player_Name']].assign(Trophy_Title=theTitle).copy()
    dfe = dfd.replace({'Year': r'[/-].*'}, {'Year': ''}, regex=True)[['Year', 'Trophy_Title', 'Player_Name']].copy()

    dff = dfe[['Year', 'Trophy_Title', 'Player_Name']].assign(Trophy_Image_URI=theUrl).copy()
    dfg = dff[['Year', 'Trophy_Title', 'Trophy_Image_URI', 'Player_Name']].copy()
    theSheetName = sport
    append_df_to_excel(dfg, 'export.xlsx', sheet_name=theSheetName)
    theFileName = str(r'export_' + sport.replace(" ", "_") + '_' + theTitle.replace(" ", "_") + '.xlsx')
    # dfg.to_excel(theFileName, sheet_name=theSheetName, index=False, header=True)
    return dfg

def getImageUrl(df,column,defaultUrl):
    imageUrl = defaultUrl
    firstRow = df[column].iloc[0]
    theType = type(firstRow)
    if (type(firstRow) == str):
        # print("first row=" + firstRow + " " + ",type=" + str(type(firstRow)))
        theFirstRow = str(firstRow)
        pattern = '^(http|https)://'
        if (re.match(pattern, theFirstRow)):
            imageUrl = firstRow
    return imageUrl

def convertCrossCountry(sport, url):
    # Cross country - boys
    r = requests.get(url)
    data = r.content
    df = pd.read_csv(BytesIO(data))
    df.columns = df.columns.to_series().apply(lambda x: x.strip())
    columns = df.columns.values
    for column in columns:
        print("column=" + column)
        if(column == 'Unnamed: 0'):
            df.rename(columns={'Unnamed: 0': 'Year'}, inplace=True)
            column = 'Year'
        if (column != 'Year'):
            imageUrl = getImageUrl(df, column, 'https://TODO')
            print("imageUrl=" + imageUrl)
            convertTable(sport, df, column, imageUrl)


convertCrossCountry('Cross Country Boys',
                    'https://docs.google.com/spreadsheets/d/16uzBqifChpbk2l8L3qx37Ij4YggH7TbDEgly67UClAg/export?format=csv&gid=0')
convertCrossCountry('Cross Country Girls',
                    'https://docs.google.com/spreadsheets/d/16uzBqifChpbk2l8L3qx37Ij4YggH7TbDEgly67UClAg/export?format=csv&gid=632751317')
